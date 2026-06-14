from urllib.parse import quote
import os, uuid, mimetypes, json
from fastapi import APIRouter, Depends, UploadFile, File, HTTPException, Form, Query
from fastapi.responses import FileResponse, JSONResponse
from sqlalchemy.orm import Session
from ..database import get_db
from ..models import Attachment, Project

UPLOAD_DIR = "/app/data/uploads"
os.makedirs(UPLOAD_DIR, exist_ok=True)
IMAGE_TYPES = {"image/jpeg", "image/png", "image/gif", "image/webp", "image/svg+xml", "image/bmp"}
MAX_SIZE = 50 * 1024 * 1024

router = APIRouter(prefix="/projects", tags=["attachments"])


def _attachment_out(a: Attachment) -> dict:
    return {
        "id": a.id, "project_id": a.project_id,
        "original_name": a.original_name,
        "file_size": a.file_size,
        "mime_type": a.mime_type,
        "is_image": a.is_image,
        "attachment_type": a.attachment_type or "general",
        "follow_up_id": a.follow_up_id,
        "created_at": a.created_at.isoformat() if a.created_at else None,
        "preview_url": f"/api/v1/presales/projects/{a.project_id}/attachments/{a.id}/file",
    }


@router.post("/{project_id}/attachments")
def upload_attachment(
    project_id: int,
    file: UploadFile = File(...),
    attachment_type: str = Form("general"),
    follow_up_id: int = Form(None),
    db: Session = Depends(get_db),
):
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(404, "项目不存在")
    content = file.file.read()
    if len(content) > MAX_SIZE:
        raise HTTPException(413, f"文件大小超过限制({MAX_SIZE // 1024 // 1024}MB)")
    ext = os.path.splitext(file.filename or "file")[1].lower()
    stored_name = f"{uuid.uuid4().hex}{ext}"
    file_path = os.path.join(UPLOAD_DIR, stored_name)
    with open(file_path, "wb") as f:
        f.write(content)
    mime_type = file.content_type or mimetypes.guess_type(file.filename)[0] or "application/octet-stream"
    is_image = mime_type in IMAGE_TYPES
    att = Attachment(
        project_id=project_id,
        original_name=file.filename or stored_name,
        stored_name=stored_name,
        file_path=file_path,
        file_size=len(content),
        mime_type=mime_type,
        is_image=is_image,
        attachment_type=attachment_type,
        follow_up_id=follow_up_id,
    )
    db.add(att)
    db.commit()
    db.refresh(att)
    return _attachment_out(att)


@router.get("/{project_id}/attachments")
def list_attachments(
    project_id: int,
    attachment_type: str = "",
    db: Session = Depends(get_db),
):
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(404, "项目不存在")
    q = db.query(Attachment).filter(Attachment.project_id == project_id)
    if attachment_type:
        q = q.filter(Attachment.attachment_type == attachment_type)
    atts = q.order_by(Attachment.created_at.desc()).all()
    return [_attachment_out(a) for a in atts]


@router.get("/{project_id}/attachments/{att_id}/file")
def get_attachment_file(project_id: int, att_id: int, db: Session = Depends(get_db)):
    att = db.query(Attachment).filter(Attachment.id == att_id).first()
    if not att:
        raise HTTPException(404, "附件不存在")
    if not os.path.exists(att.file_path):
        raise HTTPException(404, "文件已丢失")
    disposition = (
        "inline"
        if att.is_image or att.mime_type == "application/pdf"
        else "attachment"
    )
    return FileResponse(
        att.file_path,
        media_type=att.mime_type,
        filename=att.original_name,
        headers={"Content-Disposition": f"{disposition}; filename*=UTF-8''{quote(att.original_name)}"},
    )


@router.delete("/{project_id}/attachments/{att_id}")
def delete_attachment(project_id: int, att_id: int, db: Session = Depends(get_db)):
    att = db.query(Attachment).filter(Attachment.id == att_id).first()
    if not att:
        raise HTTPException(404, "附件不存在")
    if os.path.exists(att.file_path):
        os.remove(att.file_path)
    db.delete(att)
    db.commit()
    return {"ok": True}


# ── Excel 预览（多 Sheet）──

def _parse_sheet(ws, max_rows=101):
    """解析单个 worksheet 为 headers + rows。自动跳过标题行和空行，定位表头行。"""
    # 先读取所有行
    all_rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= max_rows:
            break
        all_rows.append([str(c) if c is not None else "" for c in row])

    # 找到真正的表头行：第一个有 ≥3 个非空单元格的行
    header_idx = 0
    for i, row in enumerate(all_rows):
        non_empty = [c for c in row if c.strip()]
        if len(non_empty) >= 3:
            header_idx = i
            break

    headers = all_rows[header_idx]
    # 表头之后的行作为数据
    rows = all_rows[header_idx + 1:]

    # 过滤全部为空的行
    rows = [r for r in rows if any(c.strip() for c in r)]
    # 过滤全部为空的列
    non_empty_cols = []
    for ci in range(len(headers)):
        col_has_content = headers[ci].strip() != "" or any(
            ci < len(r) and r[ci].strip() for r in rows
        )
        if col_has_content:
            non_empty_cols.append(ci)
    headers = [headers[i] for i in non_empty_cols]
    rows = [[r[i] for i in non_empty_cols if i < len(r)] for r in rows]
    return headers, rows


@router.get("/{project_id}/attachments/{att_id}/excel-preview")
def excel_preview(project_id: int, att_id: int, sheet: str = Query(""), db: Session = Depends(get_db)):
    """解析 Excel 文件，支持多 sheet。返回 sheet 列表 + 指定 sheet 表格数据"""
    att = db.query(Attachment).filter(Attachment.id == att_id).first()
    if not att:
        raise HTTPException(404, "附件不存在")
    if not os.path.exists(att.file_path):
        raise HTTPException(404, "文件已丢失")

    ext = os.path.splitext(att.original_name)[1].lower()
    if ext not in (".xlsx", ".xls", ".csv"):
        raise HTTPException(400, "仅支持 Excel/CSV 文件预览")

    try:
        import openpyxl

        wb = openpyxl.load_workbook(att.file_path, read_only=True, data_only=True)
        sheet_names = wb.sheetnames

        # 确定要读取的 sheet（默认第一个）
        target_sheet = sheet if sheet and sheet in sheet_names else sheet_names[0]
        ws = wb[target_sheet]

        headers, rows = _parse_sheet(ws)
        wb.close()

        return JSONResponse({
            "sheets": sheet_names,
            "current_sheet": target_sheet,
            "headers": headers,
            "rows": rows,
            "total_rows": len(rows),
            "truncated": ws.max_row > 101,
        })
    except ImportError:
        raise HTTPException(500, "服务器未安装 openpyxl，请联系管理员")
    except Exception as e:
        raise HTTPException(500, f"Excel 解析失败: {str(e)}")

@router.get("/{project_id}/follow-ups/{fu_id}/attachments")
def get_follow_up_attachments(project_id: int, fu_id: int, db: Session = Depends(get_db)):
    atts = db.query(Attachment).filter(
        Attachment.project_id == project_id,
        Attachment.follow_up_id == fu_id
    ).order_by(Attachment.created_at.desc()).all()
    return [_attachment_out(a) for a in atts]
